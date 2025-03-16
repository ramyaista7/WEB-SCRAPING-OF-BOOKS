import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import datetime
import os
import re
from openpyxl.chart import BarChart, Reference

# Ensure output directories exist
output_dir = r"C:\Users\Hello\Desktop\web"
image_dir = os.path.join(output_dir, "images")
os.makedirs(output_dir, exist_ok=True)
os.makedirs(image_dir, exist_ok=True)

# Generate a unique filename
filename = os.path.join(output_dir, f"books_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

books = []
max_pages = 3  # ✅ Scrape only 3 pages

# Star rating mapping
star_map = {
    "One": "⭐",
    "Two": "⭐⭐",
    "Three": "⭐⭐⭐",
    "Four": "⭐⭐⭐⭐",
    "Five": "⭐⭐⭐⭐⭐"
}

# Function to clean filenames
def clean_filename(title, max_length=50):
    return re.sub(r'[\/:*?"<>|,]', '', title)[:max_length]  # Remove invalid chars & limit length

# Web Scraping Loop
for page in range(1, max_pages + 1):
    url = f"https://books.toscrape.com/catalogue/page-{page}.html"
    response = requests.get(url)
    if response.status_code != 200:
        break

    soup = BeautifulSoup(response.content, 'html.parser')
    articles = soup.find_all('article', class_='product_pod')
    for article in articles:
        image_tag = article.find('img')
        title = " ".join(image_tag['alt'].strip().split()) if image_tag else 'NO TITLE'
        star_tag = article.find('p', class_='star-rating')
        star = star_map.get(star_tag['class'][1], "No Rating") if star_tag else 'No Rating'
        price_tag = article.find('p', class_='price_color')
        price = f"£{float(price_tag.text[1:].strip()):.2f}" if price_tag else "£0.00"


        # Download Image
        image_url = "https://books.toscrape.com/" + image_tag['src'].replace("../", "") if image_tag else None
        image_path = None
        if image_url:
            img_response = requests.get(image_url)
            if img_response.status_code == 200:
                safe_title = clean_filename(title)
                image_path = os.path.join(image_dir, f"{safe_title}.jpg")
                with open(image_path, 'wb') as img_file:
                    img_file.write(img_response.content)

        books.append([title, star, price, image_path])  # Store book details

    print(f"✅ Scraped page {page}")

# Sort books: Alphabetic first, non-alphabetic last
books = sorted(books, key=lambda x: (not x[0][0].isalpha(), x[0]))

# Create an Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Books Data"

# Headers
headers = ["SERIAL NO", "BOOK VIEW", "TITLE", "STAR RATING", "PRICE"]
ws.append(headers)

# Define Styles
header_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")  # Dark Blue
header_font = Font(bold=True, color="FFFFFF", size=18)
center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

table_font = Font(size=14, bold=True)
star_font = Font(bold=True, color="FFD700", size=16)

serial_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
title_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
price_fill = PatternFill(start_color="FFDDC1", end_color="FFDDC1", fill_type="solid")

black_border = Border(
    left=Side(border_style="thick", color="000000"),
    right=Side(border_style="thick", color="000000"),
    top=Side(border_style="thick", color="000000"),
    bottom=Side(border_style="thick", color="000000")
)

# Apply header styling
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_alignment
    cell.border = black_border

ws.row_dimensions[1].height = 80  # Bigger header row

# Add book data to sheet
row_num = 2
for i, (title, star, price, image_path) in enumerate(books, start=1):
    ws.append([i, "", title, star, price])  

    
    # ws[f"A{row_num}"].font = table_font
    ws[f"A{row_num}"].fill = serial_fill
    ws[f"A{row_num}"].font = Font(size=18, bold=True)  # Serial Number #table font
    ws[f"A{row_num}"].alignment = center_alignment

    
    # ws[f"C{row_num}"].font = table_font
    ws[f"C{row_num}"].fill = title_fill
    ws[f"C{row_num}"].alignment = center_alignment
    ws[f"C{row_num}"].font = Font(size=18, bold=True)  # Title

    # ws[f"D{row_num}"].font = star_font
    ws[f"D{row_num}"].alignment = center_alignment
    ws[f"D{row_num}"].font = Font(size=18, bold=True, color="FFD700")  # Star Rating

    
    # ws[f"E{row_num}"].font = table_font
    ws[f"E{row_num}"].fill = price_fill
    ws[f"E{row_num}"].alignment = center_alignment
    ws[f"E{row_num}"].font = Font(size=18, bold=True)  # Price

    
    ws[f"G{row_num}"].font = Font(bold=True, color="FFD700", size=18) 
    ws[f"H{row_num}"].font = Font(size=18, bold=True)


    if image_path and os.path.exists(image_path):
        img = Image(image_path)
        img.width, img.height = 120, 160  # Increased image size
        ws.row_dimensions[row_num].height = 250  # Increased row height
        ws.column_dimensions["B"].width = 40  # Increased column width for images
        ws.add_image(img, f"B{row_num}")

    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}{row_num}"].border = black_border
        ws[f"{col}{row_num}"].alignment = center_alignment

    row_num += 1

# Adjust column widths for better visibility
column_widths = {"A": 25, "B": 40, "C": 80, "D": 30, "E": 25, "G": 30, "H": 25}
for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Increase row heights for better spacing
for row in range(2, row_num):
    ws.row_dimensions[row].height = 150  # Increased height for better readability

# Star Rating Count Table
ws["G2"] = "STAR RATING"
ws["H2"] = "COUNT"
ws["G2"].font = ws["H2"].font = Font(bold=True, color="FFFFFF", size=16)
ws["G2"].fill = ws["H2"].fill = PatternFill(start_color="4B0082", end_color="4B0082", fill_type="solid")
ws["G2"].alignment = ws["H2"].alignment = center_alignment
ws["G2"].border = ws["H2"].border = black_border

star_counts = {"⭐⭐⭐⭐⭐": 0, "⭐⭐⭐⭐": 0, "⭐⭐⭐": 0, "⭐⭐": 0, "⭐": 0}
for book in books:
    if book[1] in star_counts:
        star_counts[book[1]] += 1

row = 3
for rating, count in star_counts.items():
    ws[f"G{row}"] = rating
    ws[f"G{row}"].font = Font(bold=True, color="FFD700", size=16) 
    ws[f"H{row}"] = count
    ws[f"H{row}"].fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")  # Light Yellow
    ws[f"G{row}"].border = ws[f"H{row}"].border = black_border
    ws[f"G{row}"].alignment = ws[f"H{row}"].alignment = center_alignment
    row += 1

# Add Bar Chart
chart = BarChart()
chart.title = "Star Ratings Distribution"
chart.x_axis.title = "Star Rating"
chart.y_axis.title = "Number of Books"
chart.width = 12
chart.height = 8

data = Reference(ws, min_col=8, min_row=2, max_row=row - 1)
categories = Reference(ws, min_col=7, min_row=3, max_row=row - 1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)
ws.add_chart(chart, "J5")

wb.save(filename)
print(f"\n✅ Scraping completed! Data saved as '{filename}'.")